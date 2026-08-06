import re
from io import BytesIO
from xml.etree import ElementTree

from openpyxl import load_workbook

from app.services.knowledge.normalize import normalize_knowledge_text
from app.services.knowledge.schema import (
    ParseIssue,
    ParseResult,
    RawKnowledgeUnit,
    SourceLocator,
)

_IMAGE_PATTERN = re.compile(r"!\[[^\]]*\]\([^)]*\)|<img\b[^>]*>", re.IGNORECASE)
_HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
_HEADER_VOCABULARY = {
    "分类",
    "品类",
    "问题",
    "回复",
    "话术",
    "答案",
    "回答",
    "产品名称",
    "货品名称",
}
_ANSWER_HEADERS = {"回复", "话术", "答案", "回答"}
_CATEGORY_HEADERS = {"分类", "品类"}


def parse_markdown(source_path: str, text: str) -> ParseResult:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    headings: list[tuple[int, int, str]] = []
    for line_number, line in enumerate(lines, start=1):
        match = _HEADING_PATTERN.match(line)
        if match:
            headings.append(
                (line_number, len(match.group(1)), normalize_knowledge_text(match.group(2)))
            )

    units: list[RawKnowledgeUnit] = []
    issues: list[ParseIssue] = []
    path_stack: list[str] = []
    anchors: dict[str, int] = {}
    skipped_images = len(_IMAGE_PATTERN.findall(text))
    for index, (line_number, depth, title) in enumerate(headings):
        path_stack = path_stack[: depth - 1]
        path_stack.append(title)
        path = list(path_stack)
        base_anchor = "h:" + "/".join(part.replace("/", "／") for part in path)
        anchors[base_anchor] = anchors.get(base_anchor, 0) + 1
        anchor = (
            base_anchor
            if anchors[base_anchor] == 1
            else f"{base_anchor}#{anchors[base_anchor]}"
        )
        next_line = headings[index + 1][0] if index + 1 < len(headings) else len(lines) + 1
        body = "\n".join(lines[line_number: next_line - 1])
        content = normalize_knowledge_text(_IMAGE_PATTERN.sub("", body))
        locator = SourceLocator(
            source_path=source_path,
            kind="markdown",
            anchor=anchor,
            line=line_number,
            path=path,
        )
        if not content:
            issues.append(
                ParseIssue(
                    code="empty_item",
                    severity="info",
                    message=f'Markdown section "{title}" has no text content.',
                    sources=[locator],
                )
            )
            continue
        units.append(
            RawKnowledgeUnit(
                title=title,
                content=content,
                category_path=path[:-1],
                source=locator,
            )
        )
    return ParseResult(
        units=units,
        issues=issues,
        stats={
            "sections_seen": len(headings),
            "units_emitted": len(units),
            "empty_items_skipped": len(issues),
            "skipped_images": skipped_images,
        },
    )


def parse_excel(source_path: str, content: bytes) -> ParseResult:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    units: list[RawKnowledgeUnit] = []
    issues: list[ParseIssue] = []
    rows_seen = 0
    empty_sheets = 0
    for sheet in workbook.worksheets:
        rows = [
            (row_number, [normalize_knowledge_text(value) for value in row])
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1)
            if any(normalize_knowledge_text(value) for value in row)
        ]
        if not rows:
            empty_sheets += 1
            locator = SourceLocator(
                source_path=source_path,
                kind="excel",
                anchor=f"sheet:{sheet.title}",
                sheet=sheet.title,
                path=[sheet.title],
            )
            issues.append(
                ParseIssue(
                    code="empty_sheet",
                    severity="info",
                    message=f'Excel sheet "{sheet.title}" has no cell content.',
                    sources=[locator],
                )
            )
            continue
        header = _find_header_row(rows)
        header_number, header_values = header if header else (None, [])
        question_index = _find_header_index(header_values, {"问题"})
        answer_index = _find_header_index(header_values, _ANSWER_HEADERS)
        category_index = _find_header_index(header_values, _CATEGORY_HEADERS)
        for row_number, values in rows:
            if row_number == header_number:
                continue
            rows_seen += 1
            question = _value_at(values, question_index)
            answer = _value_at(values, answer_index)
            category = _value_at(values, category_index)
            title = question or next((value for value in values if value), "")
            locator = SourceLocator(
                source_path=source_path,
                kind="excel",
                anchor=f"sheet:{sheet.title}/row:{row_number}",
                sheet=sheet.title,
                row=row_number,
                path=[sheet.title, *([category] if category else []), *([title] if title else [])],
            )
            if question_index is not None and answer_index is not None:
                if not question:
                    issues.append(
                        ParseIssue(
                            code="empty_item",
                            severity="warning",
                            message=f"Excel row {sheet.title}!{row_number} has no question.",
                            sources=[locator],
                        )
                    )
                    continue
                if not answer:
                    issues.append(
                        ParseIssue(
                            code="empty_answer",
                            severity="warning",
                            message=f"Excel row {sheet.title}!{row_number} has no answer.",
                            sources=[locator],
                        )
                    )
                    continue
                units.append(
                    RawKnowledgeUnit(
                        title=question[:80],
                        content=answer,
                        category_path=[sheet.title, *([category] if category else [])],
                        semantic_key="qa:" + "|".join(
                            part for part in [category, question] if part
                        ),
                        source=locator,
                    )
                )
                continue
            if not title:
                issues.append(
                    ParseIssue(
                        code="empty_item",
                        severity="warning",
                        message=f"Excel row {sheet.title}!{row_number} has no usable text.",
                        sources=[locator],
                    )
                )
                continue
            fields = [value for value in values if value]
            units.append(
                RawKnowledgeUnit(
                    title=title[:80],
                    content="\n".join(fields),
                    category_path=[sheet.title],
                    source=locator,
                )
            )
    workbook.close()
    return ParseResult(
        units=units,
        issues=issues,
        stats={
            "sheets_seen": len(workbook.sheetnames),
            "rows_seen": rows_seen,
            "units_emitted": len(units),
            "empty_items_skipped": len(issues) - empty_sheets,
            "empty_sheets": empty_sheets,
            "skipped_images": 0,
        },
    )


def parse_mindmap(source_path: str, text: str) -> ParseResult:
    root = ElementTree.fromstring(text)
    units: list[RawKnowledgeUnit] = []
    issues: list[ParseIssue] = []
    nodes_seen = 0
    url_references = 0

    def walk(node: ElementTree.Element, parent_path: list[str], positions: list[int]) -> None:
        nonlocal nodes_seen, url_references
        nodes_seen += 1
        content = normalize_knowledge_text(node.attrib.get("TEXT", ""))
        title = content.split("\n", 1)[0] if content else ""
        node_id = normalize_knowledge_text(node.attrib.get("ID", "")) or "path-" + "-".join(map(str, positions))
        path = [*parent_path, title] if title else list(parent_path)
        locator = SourceLocator(
            source_path=source_path,
            kind="mindmap",
            anchor=f"node:{node_id}",
            node_id=node_id,
            path=path,
        )
        if not content:
            issues.append(
                ParseIssue(
                    code="empty_item",
                    severity="info",
                    message=f'Mind-map node "{node_id}" has no text content.',
                    sources=[locator],
                )
            )
        else:
            url_references += len(re.findall(r"https?://[^\s<>\"']+", content))
            units.append(
                RawKnowledgeUnit(
                    title=title,
                    content=content,
                    category_path=parent_path,
                    source=locator,
                )
            )
        for index, child in enumerate(node.findall("node"), start=1):
            walk(child, path, [*positions, index])

    for index, node in enumerate(root.findall("node"), start=1):
        walk(node, [], [index])
    return ParseResult(
        units=units,
        issues=issues,
        stats={
            "nodes_seen": nodes_seen,
            "units_emitted": len(units),
            "empty_items_skipped": len(issues),
            "url_references": url_references,
        },
    )


def _find_header_row(rows: list[tuple[int, list[str]]]) -> tuple[int, list[str]] | None:
    candidates = [
        (sum(value in _HEADER_VOCABULARY for value in values), number, values)
        for number, values in rows[:10]
    ]
    score, number, values = max(candidates, default=(0, 0, []))
    return (number, values) if score >= 2 else None


def _find_header_index(values: list[str], accepted: set[str]) -> int | None:
    return next((index for index, value in enumerate(values) if value in accepted), None)


def _value_at(values: list[str], index: int | None) -> str:
    return values[index] if index is not None and index < len(values) else ""
